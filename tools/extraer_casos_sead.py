"""
Extrae, de cada uno de los estudios reales en `assets/estudios/*.xlsx`
(salidas de la herramienta Excel "SEAD Street Lighting Tool"), la geometria
de entrada (hoja "Input") y los resultados esperados de iluminancia (hoja
"Illuminance"), para las filas cuyo luminario ("Nombre de la luminaria") sea
un archivo .ies presente en `assets/`.

Se ignora siempre la fila "Linea base (referencia)": usa un luminario
generico precargado del Excel cuyo .ies no tenemos.

Genera `reference/casos_sead.json` (UTF-8, indentado, ensure_ascii=False).

Uso:
    PYTHONIOENCODING=utf-8 python tools/extraer_casos_sead.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
DIR_ESTUDIOS = RAIZ / "assets" / "estudios"
DIR_ASSETS = RAIZ / "assets"
SALIDA = RAIZ / "reference" / "casos_sead.json"

# Columnas (1-indexadas) de la hoja "Input", fila 3 = encabezado.
COLS_INPUT = {
    "id_simulacion": 2,       # B
    "nombre_luminaria": 3,    # C
    "tipo": 4,                # D
    "vatios": 5,              # E
    "luminarias_km": 6,       # F
    "num_carriles": 7,        # G
    "ancho_carril": 8,        # H
    "camellon": 10,           # J
    "posicion_poste": 11,     # K
    "altura_montaje": 12,     # L
    "interpostal": 13,        # M
    "retranqueo": 14,         # N
    "largo_brazo": 15,        # O
    "llf": 16,                # P
}

# Columnas (1-indexadas) de la hoja "Illuminance", fila 3 = encabezado.
COLS_ILUM = {
    "id_simulacion": 2,       # B
    "nombre_luminaria": 3,    # C
    "vatios": 5,              # E
    "promedio": 6,            # F
    "minimo": 7,              # G
    "maximo": 8,              # H
    "uniformidad": 9,         # I
    "objetivo_promedio": 10,  # J
    "objetivo_uniformidad": 11,  # K
    "cumple": 12,             # L
}

FILA_LINEA_BASE = "Línea base (referencia)"

# Inclinación del luminario (tilt), en grados, por (nombre del .xlsx, .ies).
#
# ESTE DATO NO ESTÁ EN EL ARCHIVO. La herramienta lo captura por luminario
# (`Fixtures!T36`, guardado en `FixtureData!selectedTilt`) y lo usa en el
# cálculo desde la v1.8.1, pero NO lo escribe en la hoja "Input" del reporte de
# salida: la cadena "Inclinación ( grados)" solo aparece en la hoja de
# traducciones. Así que la única forma de conocerlo es que quien corrió el
# estudio lo diga.
#
# Los dos valores de abajo los declaró el usuario y se confirmaron por
# ingeniería inversa: barriendo la inclinación, 5° y 15° reproducen las cuatro
# métricas de esa corrida al 0.002 %, mientras que 0° se desvía hasta 24 % en
# uniformidad. No es un dato de fe, es un dato medido. Ver VALIDACION.md.
#
# Lo que NO está aquí se asume 0, que es lo que el Excel usa por omisión.
INCLINACIONES = {
    ("IESResults09_02_26 12_57_51", "V1070UN2M50.ies"): 5.0,
    ("IESResults09_02_26 12_57_51", "V2100UN2M50.ies"): 15.0,
}


def _leer_filas(sh, cols: dict, fila_inicio: int = 4) -> list:
    filas = []
    r = fila_inicio
    while True:
        id_sim = sh.cell(r, cols["id_simulacion"]).value
        nombre = sh.cell(r, cols["nombre_luminaria"]).value
        if id_sim is None and nombre is None:
            # fila vacia: revisar si es solo un hueco o el final de la tabla
            fila_completa = [sh.cell(r, c).value for c in cols.values()]
            if all(v is None for v in fila_completa):
                # tolerar una fila vacia intermedia, pero cortar si la
                # siguiente tambien esta vacia
                r2 = r + 1
                fila_sig = [sh.cell(r2, c).value for c in cols.values()]
                if all(v is None for v in fila_sig):
                    break
        else:
            fila = {clave: sh.cell(r, c).value for clave, c in cols.items()}
            filas.append(fila)
        r += 1
        if r > fila_inicio + 200:  # tope de seguridad
            break
    return filas


_ANONIMOS: dict = {}


def _anonimo(ruta_xlsx: Path) -> str:
    """Identificador estable y sin datos de cliente para un estudio.

    Los .xlsx llevan el nombre de la vialidad en el nombre de archivo. El caso
    de regresion no necesita saber de que calle se trata, solo su geometria, y
    esos nombres no deben acabar en el repositorio.
    """
    clave = ruta_xlsx.stem
    if clave not in _ANONIMOS:
        _ANONIMOS[clave] = "estudio-{:02d}".format(len(_ANONIMOS) + 1)
    return _ANONIMOS[clave]


def extrae_estudio(ruta_xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)

    hojas_requeridas = {"Input", "Illuminance"}
    faltantes = hojas_requeridas - set(wb.sheetnames)
    if faltantes:
        return {
            "archivo": _anonimo(ruta_xlsx),
            "error": "faltan hojas esperadas: {}".format(sorted(faltantes)),
        }

    sh_in = wb["Input"]
    sh_il = wb["Illuminance"]

    filas_in = _leer_filas(sh_in, COLS_INPUT)
    filas_il = _leer_filas(sh_il, COLS_ILUM)

    por_id_il = {f["id_simulacion"]: f for f in filas_il}

    casos = []
    omitidos = []

    for fin in filas_in:
        id_sim = fin["id_simulacion"]
        nombre = fin["nombre_luminaria"]

        if id_sim == FILA_LINEA_BASE or nombre is None:
            omitidos.append({
                "id_simulacion": id_sim,
                "nombre_luminaria": nombre,
                "motivo": "fila de linea base (luminario generico sin .ies disponible)",
            })
            continue

        ruta_ies = DIR_ASSETS / nombre
        if not ruta_ies.exists():
            omitidos.append({
                "id_simulacion": id_sim,
                "nombre_luminaria": nombre,
                "motivo": "no existe el archivo .ies en assets/: {}".format(ruta_ies),
            })
            continue

        fil = por_id_il.get(id_sim)
        if fil is None:
            omitidos.append({
                "id_simulacion": id_sim,
                "nombre_luminaria": nombre,
                "motivo": "no se encontro fila correspondiente en la hoja Illuminance",
            })
            continue

        caso = {
            "id_simulacion": id_sim,
            "luminario": {
                "archivo_ies": nombre,
                "tipo": fin["tipo"],
                "vatios": fin["vatios"],
                "inclinacion": INCLINACIONES.get((ruta_xlsx.stem, nombre), 0.0),
            },
            "geometria": {
                "luminarias_km": fin["luminarias_km"],
                "num_carriles": fin["num_carriles"],
                "ancho_carril": fin["ancho_carril"],
                "camellon": fin["camellon"],
                "posicion_poste": (fin["posicion_poste"] or "").strip()
                if isinstance(fin["posicion_poste"], str) else fin["posicion_poste"],
                "altura_montaje": fin["altura_montaje"],
                "interpostal": fin["interpostal"],
                "retranqueo": fin["retranqueo"],
                "largo_brazo": fin["largo_brazo"],
                "llf": fin["llf"],
            },
            "esperado": {
                "promedio": fil["promedio"],
                "minimo": fil["minimo"],
                "maximo": fil["maximo"],
                "uniformidad": fil["uniformidad"],
                "objetivo_promedio": fil["objetivo_promedio"],
                "objetivo_uniformidad": fil["objetivo_uniformidad"],
                "cumple": fil["cumple"],
            },
        }
        casos.append(caso)

    return {
        "archivo": None,
        "estudio": _anonimo(ruta_xlsx),
        "casos": casos,
        "omitidos": omitidos,
    }


def main() -> None:
    rutas = sorted(DIR_ESTUDIOS.glob("*.xlsx"))
    resultado = {
        "descripcion": (
            "Casos de regresion extraidos de los estudios reales de "
            "assets/estudios/*.xlsx (SEAD Street Lighting Tool). Cada caso "
            "trae la geometria de entrada y los resultados esperados de "
            "iluminancia para un luminario cuyo .ies existe en assets/."
        ),
        "num_estudios": len(rutas),
        "estudios": [],
    }

    for ruta in rutas:
        resultado["estudios"].append(extrae_estudio(ruta))

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    with SALIDA.open("w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    total_casos = sum(len(e.get("casos", [])) for e in resultado["estudios"])
    total_omitidos = sum(len(e.get("omitidos", [])) for e in resultado["estudios"])
    print("Estudios procesados: {}".format(len(rutas)))
    print("Casos extraidos: {}".format(total_casos))
    print("Filas omitidas: {}".format(total_omitidos))
    print("Escrito: {}".format(SALIDA))


if __name__ == "__main__":
    main()
